'''
Program to build SQL insert, update, select, and delete scripts from data in an excel spreadsheet
Matt Saffert
12-31-2019
'''

import openpyxl
import tkinter
from tkinter import filedialog as tkFileDialog
import create_excel_template as template
import write_sql_scripts as write_scripts
import validate_workbook as validate
import excel_global
import sys
import os


def main():
    '''Main run function for the Excel python program. Called once on program initialization

    :return: NONE
    '''

    # try:
    # gets the mode of the program that the user would like to use
    program_mode = excel_global.getProgramMode().get()

    if program_mode == 'scripts':
        write_scripts.displayExcelFormatInstructions()  # tkinter dialog box

        output_string = "Choose the Excel workbook you'd like to make scripts for."
        excel_global.createPopUpBox(output_string)  # tkinter dialog box

        file = tkinter.Tk()
        # opens file explorer so user can choose file to read from
        file.filename = tkFileDialog.askopenfilename(
            initialdir="C:/", title="Select file to write scripts for")
        file.destroy()

        workbook = openpyxl.load_workbook(
            filename=file.filename, data_only=True)

        validate_with_sql = excel_global.createYesNoBox(
            'Would you like to validate Workbook with SQL table or generic validation?', 'SQL', 'Generic')

        write_to_sql = 'SQL'
        write_to_excel = 'Excel'
        description = 'Would you like to write the sql scripts to a ".sql" file or to an Excel spreadsheet?'
        write_to = excel_global.createYesNoBox(
            description, write_to_sql, write_to_excel)

        save_file = write_scripts.writeToExcel(workbook, validate_with_sql, write_to)

        if save_file == 'Excel':
            output_string = "Select/create the filename of Excel workbook you'd like to save/write to: "
            excel_global.createPopUpBox(
                output_string)  # tkinter dialog box

            file = tkinter.Tk()
            # opens file explorer so user can choose file to write to
            file.filename = tkFileDialog.asksaveasfilename(
                initialdir="C:/", title="Select/create file to save/write to", defaultextension=".xlsx")
            # saves new workbook with generated scripts to a user selected file
            workbook.save(file.filename)
            file.destroy()

            output_string = "Scripts saved to: '" + \
                str(file.filename) + "'"
            excel_global.createPopUpBox(
                output_string)  # tkinter dialog box
        elif save_file == '':
            output_string = "No files were changed. Closing program."
            excel_global.createPopUpBox(
                output_string)  # tkinter dialog box

    elif program_mode == 'template':
        # creates new workbook to write template to
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        template_type = template.getTemplateType()

        if template_type == 'from_table':  # generates an Excel template from a SQL database
            sql_column_names, sql_column_types, column_is_nullable, column_is_identity, sql_table_name = template.getTemplateInfo()  # tkinter dialog boxes

            worksheet.title = sql_table_name

            # allows user to select the type of script this template is for
            script_type = template.getTypeOfScriptFromUser(
                worksheet.title).get()  # tkinter dialog box

            # asks user which elements from the imported table they'd like to include in their scripts
            sql_include_row, sql_where_row, disable_include_change = template.populateClauses(
                sql_table_name, sql_column_names, column_is_nullable, column_is_identity, script_type)  # tkinter dialog boxes
        elif template_type == 'generic':  # generates a generic template with default table data
            sql_table_name, script_type, sql_column_names, sql_column_types, sql_include_row, sql_where_row, disable_include_change = template.generateGenericTemplate(
                worksheet)
            worksheet.title = sql_table_name
        else:
            excel_global.closeProgram()

        # writes the generated template to the new Excel workbook
        template.WriteTemplateToSheet(worksheet, sql_table_name, script_type,
                                      sql_column_names, sql_column_types, sql_include_row, sql_where_row, disable_include_change)

        output_string = "Select/create the filename of Excel workbook you'd like to save/write to: "
        excel_global.createPopUpBox(output_string)  # tkinter dialog box

        file = tkinter.Tk()
        # opens file explorer so user can choose file to write to
        file.filename = tkFileDialog.asksaveasfilename(
            initialdir="C:/", title="Select/create file to save/write to", defaultextension=".xlsx")
        # saves new workbook with generated template to a user selected file
        workbook.save(file.filename)
        file.destroy()

        output_string = "Excel template saved to: '" + \
            str(file.filename) + "'"
        excel_global.createPopUpBox(output_string)  # tkinter dialog box

    elif program_mode == 'validate':
        write_scripts.displayExcelFormatInstructions()  # tkinter dialog box

        output_string = "Choose the Excel workbook you'd like to validate."
        excel_global.createPopUpBox(output_string)  # tkinter dialog box

        file = tkinter.Tk()
        # opens file explorer so user can choose file to read from
        file.filename = tkFileDialog.askopenfilename(
            initialdir="C:/", title="Select file to write scripts for")
        file.destroy()

        workbook = openpyxl.load_workbook(
            filename=file.filename, data_only=True)

        validate_with_sql = excel_global.createYesNoBox(  # tkinter dialog box that asks user if they want to connect to a SQL database to validate spreadsheet
            'Would you like to validate Workbook with SQL table or generic validation?', 'SQL', 'Generic')

        any_changes = False # False if all spreadsheets fail validation
        all_sheets_okay = True # True if all spreadsheets pass validation

        for worksheet in workbook.worksheets:
            if worksheet.title != 'configuration':  # skip the configuration sheet in the Excel book
                # check if worksheet is is valid and if user wants to write scripts for them
                valid_template = validate.validWorksheet(
                    worksheet, validate_with_sql)
                all_sheets_okay = all_sheets_okay and valid_template # True if spreadsheet passes validation
                if valid_template:  # only write to Excel if the Excel spreadsheet is a valid format
                    output_string = "VALID. This worksheet will function properly with the 'Write SQL script' mode of this program."
                    excel_global.createPopUpBox(
                        output_string)  # tkinter dialog box
                    any_changes = True  # changes were made and need to be saved

        if any_changes and not all_sheets_okay:  # some but not all spreadsheets in workbook pass validation
            output_string = "CAUTION. Care must be taken building scripts with this workbook because not all sheets are in a valid form."
            excel_global.createPopUpBox(output_string)


'''
    except Exception as e:
        excel_global.createErrorBox(repr(e))
        print(repr(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        sys.exit()
'''

main()
